#!/usr/bin/env python3.8
# -*- coding: UTF-8 -*-
from openpyxl import load_workbook

workbook = load_workbook(filename='empty_book1.xlsx')
worksheet = workbook['range names']
print(worksheet.cell(row=10, column=10).value)
'''
@version:3.8
@author:chenchen
@file:read_excel.py
@time:2023/3/17 18:05
'''
