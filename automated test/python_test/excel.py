#!/usr/bin/env python3.8
# -*- coding: UTF-8 -*-
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
dest_filename = 'empty_book1.xlsx'

ws1 = wb.active
ws1.title = "range names"

for row in range(1, 40):
    ws1.append(range(600))

ws2 = wb.create_sheet(title="Pi")

ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(row=row, column=col, value="{0}".format(get_column_letter(col)))
    print(ws3.cell(row=10, column=27).value)
wb.save(filename=dest_filename)
'''
@version:3.8
@author:chenchen
@file:excel.py
@time:2023/3/17 17:00
'''
